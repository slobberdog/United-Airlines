import openpyxl as xl

path1 = 'C:\\Users\\Leona\\Documents\\GitHub\\United Airlines\\auopstats_intermediate.xlsx'
path2 = 'C:\\Users\\Leona\\Documents\\GitHub\\United Airlines\\auopstats.xlsx'

wb1 = xl.load_workbook(filename=path1)
ws1 = wb1.worksheets[1]

wb2 = xl.load_workbook(filename=path2)
ws2 = wb2.worksheets[1]

mr = ws1.max_row
mc = ws1.max_column

# copying the cell values from source 
# excel file to destination excel file
for i in range (1, mr + 1):
    for j in range (1, mc + 1):
        # reading cell value from source excel file
        c = ws1.cell(row = i, column = j)
  
        # writing the read value to destination excel file
        ws2.cell(row = i, column = j).value = c.value
  
wb2.save(str(path2))


