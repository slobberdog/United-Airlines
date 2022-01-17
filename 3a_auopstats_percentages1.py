import openpyxl as xl
from openpyxl.styles import Font, NamedStyle, Color
#import pandas as pd
from openpyxl import Workbook

path1 = 'C:\\Users\\Leona\\Documents\\GitHub\\United Airlines\\auopstats.xlsx'

template = xl.load_workbook(filename=path1)
ws = template["summaryLNB"]

ws['A4'] = "Total pilots costs (000)"
ws['A8'] = "Salaries % of total costs"
ws['A9'] = "Training % of total costs"

ws['B8'] = ('=(B5/B4)*100')
ws['B9'] = ('=(B6/B4)*100')

ws['A2'].font = Font(bold=True, name='Arial', size=10)
ws['A4'].font = Font(bold=True, name='Arial', size=10)
ws['A8'].font = Font(bold=True, name='Arial', size=10)
ws['A9'].font = Font(bold=True, name='Arial', size=10)
ws['B22'].font = Font(bold=True, name='Arial', size=9)

for i in range (2,23):

    totcosts = ws.cell(column=i,row=4).value
    percsal = ws.cell(column=i,row=5).value
    trainsal = ws.cell(column=i,row=6).value
    finalpercsal = (percsal/totcosts) * 100
    finaltrainsal = (trainsal/totcosts) * 100
    ws.cell(column=i,row=8).value = round(finalpercsal,2)
    ws.cell(column=i,row=9).value = round(finaltrainsal,2)

col_style = Font(bold=True, name='Arial', size=10)

for j in range(2,23):
    ws.cell(column=j,row=1).font = col_style

template.save(filename=path1)

